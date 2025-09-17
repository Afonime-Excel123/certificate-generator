# generator_long.py
from PIL import Image, ImageDraw, ImageFont
from openpyxl import load_workbook
from datetime import datetime
import os
import re

# -----------------------------
# CONFIG (Long Text)
# -----------------------------
import os
OUTPUT_DIR = os.path.join(os.getcwd(), "generated_certificates")
os.makedirs(OUTPUT_DIR, exist_ok=True)

# Font paths (macOS)
FONT_REGULAR = "/Library/Fonts/Georgia.ttf"
FONT_BOLD = "/Library/Fonts/Georgia Bold.ttf"

# Layout
NAME_Y = 665
WRITEUP_START_Y = 820
CERT_ID_POSITION = (900, 1274)
MAX_TEXT_WIDTH = 1600

NAME_FONT_SIZE = 100
PARAGRAPH_FONT_SIZE = 30
BOLD_PARAGRAPH_FONT_SIZE = 28
ID_FONT_SIZE = 30


def get_day_with_suffix(day):
    if 10 <= day % 100 <= 20:
        return f"{day}th"
    return f"{day}{ {1: 'st', 2: 'nd', 3: 'rd'}.get(day % 10, 'th')}"


def split_text_with_bold(text, bold_phrases):
    chunks = [(text, False)]
    for phrase in bold_phrases:
        new_chunks = []
        for chunk_text, is_bold in chunks:
            if is_bold:
                new_chunks.append((chunk_text, is_bold))
                continue
            idx = chunk_text.lower().find(phrase.lower())
            while idx != -1:
                before = chunk_text[:idx]
                match = chunk_text[idx:idx + len(phrase)]
                after = chunk_text[idx + len(phrase):]
                if before: new_chunks.append((before, False))
                new_chunks.append((match, True))
                chunk_text = after
                idx = chunk_text.lower().find(phrase.lower())
            if chunk_text: new_chunks.append((chunk_text, False))
        chunks = new_chunks
    return chunks


def wrap_text_chunks(chunks, max_width, draw, font_reg, font_bold):
    lines = []
    current_line = []
    current_width = 0
    for text, is_bold in chunks:
        words = text.split(' ')
        for word in words:
            word += " "
            clean_word = word.replace('\n', '')
            font = font_bold if is_bold else font_reg
            word_width = draw.textlength(clean_word, font=font)
            if current_width + word_width <= max_width:
                current_line.append((clean_word, font))
                current_width += word_width
            else:
                lines.append(current_line)
                current_line = [(clean_word, font)]
                current_width = word_width
    if current_line: lines.append(current_line)
    return lines


def generate_certificates(excel_path, template_path):
    try:
        wb = load_workbook(excel_path)
        sheet = wb.active
    except Exception as e:
        return {"error": f"Failed to read Excel: {e}"}

    try:
        name_font = ImageFont.truetype(FONT_REGULAR, NAME_FONT_SIZE)
        paragraph_font = ImageFont.truetype(FONT_REGULAR, PARAGRAPH_FONT_SIZE)
        bold_paragraph_font = ImageFont.truetype(FONT_BOLD, BOLD_PARAGRAPH_FONT_SIZE)
        id_font = ImageFont.truetype(FONT_REGULAR, ID_FONT_SIZE)
    except Exception as e:
        return {"error": f"Font error: {e}. Try using Arial."}

    results = []
    errors = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        name = str(row[0]).strip() if row[0] else ""
        raw_date = row[1]
        writeup_template = str(row[2]).strip() if row[2] else ""
        cert_id = str(row[3]).strip() if row[3] else ""
        course_title = str(row[4]).strip() if row[4] else ""

        if not name or not course_title:
            continue

        # Format Date
        try:
            if isinstance(raw_date, datetime):
                formatted_date = f"{get_day_with_suffix(raw_date.day)} of {raw_date.strftime('%B')}"
            elif isinstance(raw_date, (int, float)):
                excel_date = datetime.fromordinal(datetime(1900, 1, 1).toordinal() + int(raw_date) - 2)
                formatted_date = f"{get_day_with_suffix(excel_date.day)} of {excel_date.strftime('%B')}"
            elif isinstance(raw_date, str):
                try:
                    parsed = datetime.strptime(raw_date.strip(), "%d/%m/%Y")
                    formatted_date = f"{get_day_with_suffix(parsed.day)} of {parsed.strftime('%B')}"
                except:
                    formatted_date = raw_date.strip()
            else:
                formatted_date = str(raw_date)
        except Exception as e:
            formatted_date = str(raw_date)
            errors.append(f"Date error for {name}: {e}")

        # Load Template
        try:
            template_image = Image.open(template_path)
        except Exception as e:
            return {"error": f"Failed to load template: {e}"}

        cert = template_image.copy()
        draw = ImageDraw.Draw(cert)

        # Center Name (Perfect center)
        try:
            name_bbox = draw.textbbox((0, 0), name, font=name_font)
            name_width = name_bbox[2] - name_bbox[0]
            name_x = (template_image.width - name_width) // 2
            draw.text((name_x, NAME_Y), name, font=name_font, fill="black")
        except Exception as e:
            errors.append(f"Name draw error for {name}: {e}")
            continue

        # Prepare Write-up
        full_text = writeup_template.replace("{Course}", "{course}").replace("{Date}", "{date}")
        try:
            full_text = full_text.format(course=course_title, date=formatted_date)
            full_text = re.sub(r'\s+([.,:;!?])', r'\1', full_text)
        except Exception as e:
            full_text = f"has successfully completed the {course_title} course on {formatted_date}."
            errors.append(f"Write-up format error for {name}: {e}")

        # Split & Wrap
        bold_phrases = [course_title, formatted_date]
        text_chunks = split_text_with_bold(full_text, bold_phrases)
        wrapped_lines = wrap_text_chunks(
            text_chunks, MAX_TEXT_WIDTH, draw, paragraph_font, bold_paragraph_font
        )

        # Draw Write-up
        y = WRITEUP_START_Y
        line_height = paragraph_font.getbbox("Ay")[3] + 10
        for line in wrapped_lines:
            line_width = sum(draw.textlength(word.replace('\n', ''), font=font) for word, font in line)
            x = (template_image.width - line_width) // 2
            for word, font in line:
                clean_word = word.replace('\n', '')
                draw.text((x, y), clean_word, font=font, fill="black")
                x += draw.textlength(clean_word, font=font)
            y += line_height

        # Certificate ID
        draw.text(CERT_ID_POSITION, f"Certificate ID: {cert_id}", font=id_font, fill="black")

        # Save
        safe_name = "".join(c for c in f"{name}_{course_title}" if c.isalnum() or c in " _-").replace(" ", "_")
        output_filename = f"{safe_name}_certificate.pdf"
        output_path = os.path.join(OUTPUT_DIR, output_filename)
        cert.convert("RGB").save(output_path, "PDF", resolution=100.0)
        results.append(output_filename)

    wb.close()
    return {"success": True, "generated": results, "errors": errors}